##########################################
# ExcelManager.py
# Author: Ethan H. Eastwood
# Date Last Updated: 2/9/2024
##########################################

import openpyxl as xl

global wb
global path

##########################################
#Attempts to open the workbook from __document_file_path__.
#Returns None if not found.
##########################################
def OpenWorkbook(file_path):
    try:
        global path
        global wb

        path = file_path
        wb = xl.load_workbook(path) 

        return wb
    
    except:
        print("Failed to open workbook. Does file exist?")
        return None
    
##########################################
#Attempts to open a worksheet by name for the specified workbook.
#Returns None if not found.
##########################################
def OpenSpecificSheet(ws_name):
    
    try:
        return wb[ws_name]
    except:
        print("Worksheet was not found. Does sheet already exist?")
        return None

##########################################
#Appends the data from provided list to the excel document.
#Data should be a list of lists.
##########################################    
def LoadDataToSheet(ws, dataList):

    if ws.max_row == 0:
        headers = list(dataList[0].keys())
        for col, header in enumerate(headers, start=1):
            ws.cell(row = 1, column=col, value=header)

    for row_data in dataList:
        values = list(row_data.values())
        ws.append(values)

    SaveDataToWorkbook()

##########################################
#Saves current open workbook to file path
##########################################  
def SaveDataToWorkbook():
    wb.save(path)

##########################################
#Formats a specified column in the sheet to be an Integer value
#Use AFTER loading data to sheet.
########################################## 
def ConvertColumnToInt(ws, col_num):
    for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
        for cell in row:
            try:
                cell.value = int(cell.value)
            except (ValueError, TypeError):
                pass

    SaveDataToWorkbook()

##########################################
#Formats a specified column in the sheet to be a Float value
#Use AFTER loading data to sheet.
##########################################       
def ConvertColumnToFloat(ws, col_num):
    for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
        for cell in row:
            try:
                cell.value = float(cell.value)
            except (ValueError, TypeError):
                pass

    SaveDataToWorkbook()
